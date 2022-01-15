import pyautogui                # automation library
import time                     # time library
import pyperclip                # copy paste allowing library
from openpyxl import Workbook, load_workbook # excel library

# Global variables
# pyautogui.PAUSE = 2.5 # Set up a pause after each PyAutoGUI call
excel_file = load_workbook('facebook_groups.xlsx')
excel_sheet = excel_file['automatizavimo_grupes']
post_link = "https://fb.watch/aydyGVG3JQ/"
content = '🌿 "Sveika, skanu, sotu, o svarbiausia patogu ir greita" - Sigita"'
cta1 = 'Išsirink savo glotnutį 👉'
cta2 = 'https://smutifruti.lt/'

def main():
   # some time to prepare the browser
   print("3 Seconds to prepare the browser")
   for i in range(4):
      time.sleep(1)
      print("Prepare browser" + " " + str(i) + "/3")
   time.sleep(1)
   print("Opening browser")
   time.sleep(1)
   print("\n")
   open_tab();
   count = 0
   scriptoPradzia = time.time()
   for row in excel_sheet.iter_rows():
      postoPradzia = time.time()
      group_url = row[1].value  # fetch group id from excel
      group_name = row[0].value # fetch group name from excel
      link = 'https://facebook.com/groups/'+str(group_url) # pro
      # type group name
      time.sleep(3)
      pyperclip.copy(link)
      pyautogui.hotkey('ctrl', 'v') # paste
      pyautogui.typewrite('\n')
      print("Atsidariau" + " " + str(group_name))
      print("\n")
      # let browser window load
      print("5 seconds to let the browser window load")
      for i in range(6):
         time.sleep(1)
         print("Browser window load" + " " + str(i) + "/5")
      open_writing_window();
      write_content();
      prepare_next();
      print("Done with" + " " + str(group_name)+"."+" " + "It took {0} seconds" .format(time.time() - postoPradzia))
      print("--------------------------------------------------------------------------------------")
      print("\n")       # new line
      count +=1 # variable will increment every loop iteration
   print("Papostinau i" + " " + str(count) + " " + "grupes.") # how many groups I have posted to 
   print("It took {0} seconds" .format(time.time() - scriptoPradzia)) # how long it took for the script to run
      
def open_tab():
    # opens new tab in chrome
    pyautogui.keyDown('ctrl')
    pyautogui.keyDown('t')
    pyautogui.keyUp('t')
    pyautogui.keyUp('ctrl')

def open_writing_window():
   try:
      x, y = pyautogui.locateCenterOnScreen("/home/arvydas/Dropbox/projects/facebook_automated_groups/resources/create_public_post.png", region=(1069, 673, 400, 400))
      print("The image 'create_public_post.png' was found.")
      pyautogui.click(x,y)
   except TypeError:
      print("Could not locate the image - Create a public post...")
      a, b = pyautogui.locateCenterOnScreen("/home/arvydas/Dropbox/projects/facebook_automated_groups/resources/write_something.png", region=(1069, 673, 400, 400))
      print("The image 'write something' was found")
      pyautogui.click(a,b)
   time.sleep(2)
      
def write_content():
   pyperclip.copy(post_link)
   time.sleep(1)
   pyautogui.hotkey('ctrl', 'v') # paste
   pyautogui.hotkey('ctrl','a') # select all
   pyautogui.press('backspace') # link not necessary anymore, delete
   pyperclip.copy(content)
   pyautogui.hotkey('ctrl', 'v') # paste
   pyautogui.press('enter')      # newline
   pyperclip.copy(cta1)
   pyautogui.hotkey('ctrl', 'v') # paste
   pyautogui.press('enter')      # newline
   pyperclip.copy(cta2)
   pyautogui.hotkey('ctrl', 'v') # paste
   time.sleep(1)
    
   # half(left) screen Acer Aspire V3 771G
   pyautogui.click(1666, 525) # turn off url link according to your screen pixel location (xdotool on linux)
   time.sleep(1)
   pyautogui.click(1448, 950) # Click POST according to your screen pixel location (xdotool on linux)
    
    # # whole screen Lenovo x61s
    # pyautogui.click(751, 478) # turn off url link according to your screen pixel location (xdotool on linux)
    # pyautogui.click(533, 666) # Click POST according to your screen pixel location (xdotool on linux)

def prepare_next():
    pyautogui.write(['f6'])     # mark search field, prepare for new link input
      
if __name__ == '__main__':
	main()
